import os
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

class ReportGenerator:
    """PDF・Excelレポート出力ロジック"""

    @staticmethod
    def export_to_excel(filepath: str, results: List[Dict[str, Any]], metadata: Dict[str, Any]):
        """シミュレーション結果をExcelに出力します。"""
        wb = Workbook()
        
        # 1. キャッシュフローシート
        ws = wb.active
        ws.title = "キャッシュフロー推移"
        
        # タイトル行
        ws.append([f"LifeCanvas ライフプランレポート: {metadata.get('name', 'シミュレーション')}"])
        ws.merge_cells("A1:G1")
        ws.cell(1, 1).font = Font(name="Meiryo", size=16, bold=True)
        ws.append([]) # 空白行
        
        # ヘッダー
        headers = [
            "年数", "年齢(夫)", "年齢(妻)", "年間収入(手取り)", 
            "生活費支出", "教育費支出", "住宅費支出", "マイカー費用", "保険料",
            "年間収支", "現預金残高", "運用資産残高", "純金融資産高"
        ]
        ws.append(headers)
        
        # スタイルの定義
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        header_font = Font(name="Meiryo", size=11, bold=True)
        data_font = Font(name="Meiryo", size=10)
        
        # ヘッダー行の装飾
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(3, col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
        # データ行
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        for r_data in results:
            h_age = r_data.get("husband_age", "-")
            w_age = r_data.get("wife_age", "-")
            
            # 万円単位に変換（小数点第2位まで）
            row = [
                r_data["year"],
                h_age,
                w_age,
                round(r_data["inflow"] / 10000.0, 1),
                round(r_data["living_cost"] / 10000.0, 1),
                round(r_data["education_cost"] / 10000.0, 1),
                round(r_data["housing_cost"] / 10000.0, 1),
                round(r_data["car_cost"] / 10000.0, 1),
                round(r_data["insurance_cost"] / 10000.0, 1),
                round(r_data["net_cash_flow"] / 10000.0, 1),
                round(r_data["cash_balance"] / 10000.0, 1),
                round(r_data["investment_balance"] / 10000.0, 1),
                round((r_data["cash_balance"] + r_data["investment_balance"]) / 10000.0, 1)
            ]
            ws.append(row)
            
            # ボーダー適用
            curr_row = ws.max_row
            for col_idx in range(1, len(row) + 1):
                c = ws.cell(curr_row, col_idx)
                c.font = data_font
                c.border = thin_border
                if col_idx >= 4:
                    c.number_format = '#,##0.0'
                    
        from openpyxl.utils import get_column_letter

        # 列幅の自動調整
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        wb.save(filepath)

    @staticmethod
    def export_timeline_to_excel(filepath: str, results: List[Dict[str, Any]], events: List[Any], metadata: Dict[str, Any]):
        """タイムライン・イベントのマトリックスをExcelに出力します。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "ライフイベント年表"
        
        # タイトル
        ws.append([f"LifeCanvas ライフイベント年表: {metadata.get('name', 'シミュレーション')}"])
        ws.merge_cells("A1:H1")
        ws.cell(1, 1).font = Font(name="Meiryo", size=16, bold=True)
        ws.append([]) # 空白行
        
        # ヘッダー
        headers = [
            "西暦", "年後", "夫 (年齢/年収)", "妻 (年齢/年収)", 
            "子供", "発生イベント", "資金・キャッシュフロー (万円)", "純資産 (万円)"
        ]
        ws.append(headers)
        
        # ヘッダースタイル
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(name="Meiryo", size=11, bold=True)
        valign_center = Alignment(vertical="center")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(3, col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        data_font = Font(name="Meiryo", size=10)
        event_font = Font(name="Meiryo", size=10, color="FF0000", bold=True)
        
        curr_row = 4
        for y, dat in enumerate(results):
            current_year = 2026 + y
            
            # 夫
            h_age = dat.get("husband_age", "-")
            h_inc = dat.get("husband_gross_income", 0) / 10000
            h_pen = dat.get("husband_pension", 0) / 10000
            h_ben = dat.get("husband_benefit", 0) / 10000
            if h_pen > 0 and h_inc == 0:
                h_text = f"{h_age}歳 / 年金{h_pen:.0f}万"
            else:
                h_text = f"{h_age}歳 / 年収{h_inc:.0f}万" if h_age != "-" else "-"
                if h_ben > 0:
                    h_text += f"\n(給付{h_ben:.0f}万)"
                
            # 妻
            w_age = dat.get("wife_age", "-")
            w_inc = dat.get("wife_gross_income", 0) / 10000
            w_pen = dat.get("wife_pension", 0) / 10000
            w_ben = dat.get("wife_benefit", 0) / 10000
            if w_pen > 0 and w_inc == 0:
                w_text = f"{w_age}歳 / 年金{w_pen:.0f}万"
            else:
                w_text = f"{w_age}歳 / 年収{w_inc:.0f}万" if w_age != "-" else "-"
                if w_ben > 0:
                    w_text += f"\n(給付{w_ben:.0f}万)"
                
            # 子供
            c_ages = dat.get("children_ages", [])
            c_text = ", ".join(c_ages) if c_ages else "-"
            
            # イベント
            year_events = [e for e in events if e.elapsed_year == y]
            system_events = dat.get("system_events", [])
            
            ev_str_list = []
            for e in year_events:
                ev_str_list.append(f"■ {e.name}")
            for se in system_events:
                ev_str_list.append(f"◇ {se}")
                
            ev_text = "\n".join(ev_str_list) if ev_str_list else "-"
            
            # 資金フロー
            inc = dat.get("inflow", 0) / 10000
            out = dat.get("outflow", 0) / 10000
            ncf = dat.get("net_cash_flow", 0) / 10000
            inv_dep = dat.get("investment_deposit", 0) / 10000
            inv_sold = dat.get("investment_sold", 0) / 10000
            cash = dat.get("cash_balance", 0) / 10000
            flow_text = f"通常収入: {inc:,.0f}万\n通常支出: {out:,.0f}万\n通常収支: {ncf:+,.0f}万\nNISA拠出: -{inv_dep:,.0f}万\nNISA売却: +{inv_sold:,.0f}万\n最終現預金: {cash:,.0f}万"
            nw = dat.get("net_worth", 0) / 10000
            
            row = [
                f"{current_year}年",
                f"{y}年後",
                h_text,
                w_text,
                c_text,
                ev_text,
                flow_text,
                round(nw, 1)
            ]
            ws.append(row)
            
            # スタイル適用
            for col_idx in range(1, len(row) + 1):
                c = ws.cell(curr_row, col_idx)
                c.border = thin_border
                c.alignment = valign_center
                if col_idx == 6 and ev_str_list:
                    c.font = event_font
                    c.alignment = Alignment(wrap_text=True, vertical="center") # セル内改行
                else:
                    c.font = data_font
                if col_idx == 8:
                    c.number_format = '#,##0.0'
            curr_row += 1

        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.row < 3: continue # タイトル無視
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                max_line = max([len(l)*1.5 for l in lines] + [0])
                max_len = max(max_len, max_line)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len, 10)

        wb.save(filepath)

    @staticmethod
    def export_to_pdf(filepath: str, results: List[Dict[str, Any]], metadata: Dict[str, Any]):
        """シミュレーション結果をPDFに出力します。"""
        # A4 用紙に適用
        doc = SimpleDocTemplate(filepath, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []
        
        # reportlabで日本語フォントを使用するためにシステムフォント（MS UI GothicやHelvetica-Bold）を使用
        # WindowsのMS Gothicなどを登録
        msgothic_path = r"C:\Windows\Fonts\msgothic.ttc"
        if os.path.exists(msgothic_path):
            pdfmetrics.registerFont(TTFont("MSGothic", msgothic_path))
            font_name = "MSGothic"
        else:
            # 標準フォントで代用（英語フォントベース）
            font_name = "Helvetica"

        # スタイルの設定
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name="TitleStyle",
            parent=styles["Heading1"],
            fontName=font_name,
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#333333"),
            alignment=0 # 左寄せ
        )
        
        body_style = ParagraphStyle(
            name="BodyStyle",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#555555")
        )
        
        th_style = ParagraphStyle(
            name="TableHeader",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=9,
            leading=11,
            bold=True,
            textColor=colors.white,
            alignment=1 # センター
        )

        tr_style = ParagraphStyle(
            name="TableRow",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=8,
            leading=10,
            alignment=1
        )

        # 1. タイトル
        title_p = Paragraph(f"🎨 LifeCanvas ライフプランレポート: {metadata.get('name', 'プラン')}", title_style)
        story.append(title_p)
        story.append(Spacer(1, 15))

        # 2. 基本情報・解説
        intro_text = (
            "本シミュレーションは、ご指定された「家族構成」「ライフイベント」「住宅」「車」「投資」を元に算出された"
            "40年間の長期財務キャッシュフロー予測です。キャッシュフロー上の手元現金が0円を下回らないか、"
            "また教育費のピークや老後の資金が安全に確保されているかを客観的に確認することができます。"
        )
        story.append(Paragraph(intro_text, body_style))
        story.append(Spacer(1, 20))

        # 3. キャッシュフロー指標テーブル (主な経過年数に間引いて表示 A4の1ページに収めるため)
        # 0, 5, 10, 15, 20, 25, 30, 35, 39年目を表示
        selected_years = [0, 5, 10, 15, 20, 25, 30, 35, 39]
        
        table_data = [[
            Paragraph("経過年", th_style),
            Paragraph("夫年齢", th_style),
            Paragraph("妻年齢", th_style),
            Paragraph("手取り年収", th_style),
            Paragraph("教育費", th_style),
            Paragraph("住宅関連費", th_style),
            Paragraph("年間収支", th_style),
            Paragraph("現預金高", th_style),
            Paragraph("運用資産高", th_style)
        ]]

        for yr in selected_years:
            if yr < len(results):
                r = results[yr]
                h_age = str(r.get("husband_age", "-"))
                w_age = str(r.get("wife_age", "-"))
                
                table_data.append([
                    Paragraph(f"{r['year']}年目", tr_style),
                    Paragraph(f"{h_age}歳", tr_style),
                    Paragraph(f"{w_age}歳", tr_style),
                    Paragraph(f"{r['husband_net_income']+r['wife_net_income']:.0f}円", tr_style),
                    Paragraph(f"{r['education_cost']:.0f}円", tr_style),
                    Paragraph(f"{r['housing_cost']:.0f}円", tr_style),
                    Paragraph(f"{r['net_cash_flow']:.0f}円", tr_style),
                    Paragraph(f"{r['cash_balance']:.0f}円", tr_style),
                    Paragraph(f"{r['investment_balance']:.0f}円", tr_style)
                ])

        # テーブルスタイルの設定
        col_widths = [45, 45, 45, 65, 55, 60, 60, 65, 65]
        t = Table(table_data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2383e2")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        story.append(t)
        story.append(Spacer(1, 20))

        # 4. AI診断アドバイス（最終サマリー）
        story.append(Paragraph("<b>📈 AIインサイト診断</b>", body_style))
        story.append(Spacer(1, 5))
        
        # 簡易診断文
        final_assets = (results[-1]["cash_balance"] + results[-1]["investment_balance"]) if results else 0
        if final_assets > 0:
            advice = "シミュレーションの結果、80歳時点での金融純資産は黒字を維持しています。現状の収支バランスは良好に保たれています。NISA等の積立複利効果が十分に寄与しています。"
        else:
            advice = "シミュレーションの結果、将来的にキャッシュフローが赤字に転落し、資産が不足する年があります。生活費の最適化、または住宅ローンなどの固定費支払い時期の調整をお勧めします。"
        
        story.append(Paragraph(advice, body_style))

        # 文書ビルド
        doc.build(story)
